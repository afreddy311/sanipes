from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Border, Side, PatternFill
import datetime
from openpyxl.drawing.image import Image

# Ruta del archivo de imagen que deseas agregar a las nuevas hojas
imagen_ruta = 'C:/dev/Filusi/sanipes/imagenes/Min de la prod.png'
imagen_ruta2 = 'C:/dev/Filusi/sanipes/imagenes/Sanipes.png'

# Cargar el archivo existente y la plantilla
archivo_origen = load_workbook('C:/dev/Filusi/sanipes/excels/03 MARZO 2025.xlsx')
hoja_origen = archivo_origen['INGRESOS SERIE 001']

archivo_destino = load_workbook('C:/dev/Filusi/sanipes/excels/PLANTILLA.xlsx')
hoja_plantilla = archivo_destino['923']


# Función para calcular la posición y tamaño de la imagen
def ajustar_imagen(imagen, celda_inicio, celda_fin, hoja):
    # Obtener las coordenadas de las celdas
    inicio = hoja[celda_inicio]
    fin = hoja[celda_fin]

    # Calcular el rango (ancho y alto)
    ancho = fin.column - inicio.column
    alto = fin.row - inicio.row

    # Ajustar la imagen según el rango (puedes experimentar con los valores escala)
    imagen.width = ancho*70  # Multiplicador para ajustar ancho (ajusta según necesidad)
    imagen.height = alto*33.7  # Multiplicador para ajustar alto (ajusta según necesidad)

    # Ubicar la imagen en la celda de inicio
    imagen.anchor = celda_inicio

def ajustar_imagen2(imagen, celda_inicio, celda_fin, hoja):
    # Obtener las coordenadas de las celdas
    inicio = hoja[celda_inicio]
    fin = hoja[celda_fin]

    # Calcular el rango (ancho y alto)
    ancho = fin.column - inicio.column
    alto = fin.row - inicio.row

    # Ajustar la imagen según el rango (puedes experimentar con los valores escala)
    imagen.width = ancho*120  # Multiplicador para ajustar ancho (ajusta según necesidad)
    imagen.height = alto*23  # Multiplicador para ajustar alto (ajusta según necesidad)

    # Ubicar la imagen en la celda de inicio
    imagen.anchor = celda_inicio


def desglosar_fecha(fecha_str):
    fecha = datetime.datetime.strptime(fecha_str, "%d.%m.%Y")
    return fecha.day, fecha.month, fecha.year

# Función para copiar valores y estilos (con fondo blanco por defecto)
def copiar_estilos_plantilla(celda_destino, celda_plantilla):
    # Copiar fondo blanco si existe o usar blanco por defecto
    if celda_plantilla.fill and celda_plantilla.fill.fill_type:
        celda_destino.fill = copy(celda_plantilla.fill)
    else:
        celda_destino.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Copiar bordes en negrita solamente si existen
    if celda_plantilla.border and any([
        celda_plantilla.border.left.style,
        celda_plantilla.border.right.style,
        celda_plantilla.border.top.style,
        celda_plantilla.border.bottom.style
    ]):
        celda_destino.border = copy(celda_plantilla.border)
    else:
        # Borde vacío (sin bordes)
        celda_destino.border = Border(
            left=Side(style=None),
            right=Side(style=None),
            top=Side(style=None),
            bottom=Side(style=None)
        )

# Crear nuevas hojas basadas en datos de la columna especificada
columna_valores = hoja_origen['H']
for celda in columna_valores:
    if celda.value and isinstance(celda.value, (int, float)):
        valor_celda = str(celda.value)

        # Crear una nueva hoja copiando la plantilla
        hoja_destino = archivo_destino.copy_worksheet(hoja_plantilla)
        hoja_destino.title = valor_celda

        # Copiar estilos desde la plantilla a la hoja nueva
        for fila_plantilla in hoja_plantilla.iter_rows():
            for celda_plantilla in fila_plantilla:
                nueva_celda = hoja_destino.cell(row=celda_plantilla.row, column=celda_plantilla.column)
                copiar_estilos_plantilla(nueva_celda, celda_plantilla)

        # Agregar datos desde la hoja de origen a posiciones específicas
        fila = hoja_origen[celda.row]
        fecha_str = fila[0].value
        if fecha_str:
            dia, mes, año = desglosar_fecha(fecha_str)
            hoja_destino['M9'] = dia
            hoja_destino['N9'] = mes
            hoja_destino['O9'] = año

        hoja_destino['H22'] = fila[1].value
        hoja_destino['H24'] = fila[2].value
        hoja_destino['H26'] = fila[3].value
        hoja_destino['N7'] = fila[7].value
        hoja_destino['G28'] = fila[8].value

        
        # Agregar la imagen a la nueva hoja
        nueva_imagen = Image(imagen_ruta)       
        ajustar_imagen(nueva_imagen, 'B2', 'F4', hoja_destino)  # Por ejemplo, desde A1 hasta C5
        hoja_destino.add_image(nueva_imagen, 'B2')  #  celda donde deseas colocar la imagen

        # Agregar la imagen a la nueva hoja
        nueva_imagen = Image(imagen_ruta2)       
        ajustar_imagen2(nueva_imagen, 'L2', 'N5', hoja_destino)  # Por ejemplo, desde A1 hasta C5
        hoja_destino.add_image(nueva_imagen, 'L2')  #  celda donde deseas colocar la imagen
    
# Eliminar la hoja de plantilla después de crear nuevas hojas
del archivo_destino['923']

# Guardar el archivo de destino
archivo_destino.save('MARZO SERIE 001.xlsx')

print("Hojas creadas y valores copiados exitosamente.")